import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from io import BytesIO

st.set_page_config(page_title="QuickBooks Budget Builder", layout="wide")

# YMCA Blue ribbon header
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@300;600;700&display=swap');
    
    .ymca-header {
        background: #43A5E6;
        padding: 0.6rem 2rem;
        margin: -1rem -1rem 1.5rem -1rem;
        border-radius: 0;
    }
    .ymca-title {
        color: white;
        font-size: 1.1rem;
        font-weight: 700;
        letter-spacing: 0.5px;
        margin: 0;
        font-family: 'Montserrat', 'Helvetica Neue', Helvetica, Arial, sans-serif;
        text-transform: uppercase;
    }
    .main-title {
        color: #2c3e50;
        font-size: 2rem;
        font-weight: 700;
        margin: 1.5rem 0 0.5rem 0;
        font-family: 'Montserrat', -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    }
    </style>
    <div class="ymca-header">
        <div class="ymca-title">Skagit Valley Family YMCA</div>
    </div>
    <div class="main-title">QuickBooks Budget File Generator</div>
""", unsafe_allow_html=True)

# File uploaders
col1, col2 = st.columns(2)
with col1:
    intersections_file = st.file_uploader("Upload Intersections File", type=['csv'])
with col2:
    hierarchies_file = st.file_uploader("Upload Hierarchies File", type=['csv'])

if intersections_file and hierarchies_file:
    
    # Load data
    with st.spinner("Loading data..."):
        try:
            intersections = pd.read_csv(intersections_file, low_memory=False)
            hierarchies = pd.read_csv(hierarchies_file)
        except Exception as e:
            st.error(f"❌ Error loading files: {str(e)}")
            st.stop()
    
    # Validate hierarchies
    required_hier_cols = ['_dim', '_member_name', '_member_alias', '_parent_name']
    if not all(col in hierarchies.columns for col in required_hier_cols):
        st.error("❌ **Hierarchies file is incorrect!** Please upload the correct file.")
        st.stop()
    
    # Build P&L account list
    account_hier = hierarchies[hierarchies['_dim'] == 'Account'].copy()
    
    def get_descendants(parent_name, hier_df):
        descendants = set()
        children = hier_df[hier_df['_parent_name'] == parent_name]['_member_name'].tolist()
        for child in children:
            descendants.add(child)
            descendants.update(get_descendants(child, hier_df))
        return descendants
    
    pl_accounts = get_descendants('Net Income', account_hier)
    
    # Get available scenarios and years
    available_scenarios = sorted([s for s in intersections['_Scenario'].unique() if pd.notna(s)])
    available_years = sorted(list(set([int(y) for y in intersections['_Year'].unique() 
                                        if pd.notna(y) and str(y) != 'Undefined'])))
    
    st.success(f"✓ Loaded {len(intersections):,} records - Found {len(available_scenarios)} scenarios and {len(available_years)} years")
    
    # Filter Configuration
    st.subheader("Filter Configuration")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.info(f"🔒 **Net Income (P&L Only)**\n\nFiltering to {len(pl_accounts)} P&L accounts")
    
    with col2:
        selected_scenario = st.selectbox("Scenario", available_scenarios)
    
    with col3:
        selected_year = st.selectbox("Year", available_years, index=0)  # Start with first year (2023)
    
    # QuickBooks Configuration
    st.subheader("QuickBooks Configuration")
    col1, col2, col3 = st.columns(3)
    
    with col1:
        company_name = st.text_input("Company Name", value="Skagit Valley Family YMCA")
    with col2:
        budget_name = st.text_input("Budget Name", value=f"{selected_scenario}_FY{str(selected_year)[2:]}_P&L")
    with col3:
        fiscal_year = st.text_input("Fiscal Period", value=f"FY {selected_year} (Jan {selected_year} - Dec {selected_year})")
    
    # Custom styling
    st.markdown("""
        <style>
        .stButton > button[kind="primary"] {
            background-color: #4B5FAA;
            color: white;
            border: none;
            border-radius: 8px;
            font-weight: 400;
            letter-spacing: 0.3px;
            transition: all 0.2s ease;
        }
        .stButton > button[kind="primary"]:hover {
            background-color: #3D4E8F;
            transform: translateY(-1px);
            box-shadow: 0 4px 12px rgba(75, 95, 170, 0.3);
        }
        </style>
    """, unsafe_allow_html=True)
    
    if st.button("▸ Generate QuickBooks Budget File", type="primary"):
        
        try:
            with st.spinner("Processing data..."):
                
                # Filter data (convert year to string for comparison)
                df = intersections[
                    (intersections['_Account'].astype(str).isin(pl_accounts)) &
                    (intersections['_Year'] == str(selected_year)) &
                    (intersections['_Scenario'] == selected_scenario)
                ].copy()
                
                st.info(f"▸ After filtering: {len(df):,} records")
                
                # Convert periods to month names
                def period_to_month(period):
                    try:
                        period_int = int(period)
                        return {
                            1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun',
                            7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'
                        }.get(period_int, None)
                    except (ValueError, TypeError):
                        return None
                
                df['Period_Name'] = df['_Period'].apply(period_to_month)
                df = df[df['Period_Name'].notna()].copy()
                
                # Convert values to numeric
                df['_value'] = pd.to_numeric(df['_value'], errors='coerce')
                df = df[df['_value'].notna()].copy()
                
                st.info(f"▸ After data cleaning: {len(df):,} valid records")
                
                if len(df) == 0:
                    st.error(f"❌ No data found for {selected_scenario} {selected_year}")
                    st.stop()
                
                # Build account lookup
                account_lookup = {}
                for _, row in account_hier.iterrows():
                    code = row['_member_name']
                    alias = row['_member_alias'] if pd.notna(row['_member_alias']) else code
                    account_lookup[code] = {
                        'name': code,
                        'alias': alias,
                        'parent': row['_parent_name'] if pd.notna(row['_parent_name']) else None
                    }
                
                def get_account_level(account_code, lookup):
                    level = 0
                    current = str(account_code)
                    while current and current in lookup and lookup[current]['parent']:
                        level += 1
                        current = lookup[current]['parent']
                        if level > 10:
                            break
                    return level
                
                def format_account_name(account_code, lookup):
                    if isinstance(account_code, (int, float)):
                        account_str = str(int(account_code))
                    else:
                        account_str = str(account_code)
                    
                    if account_str not in lookup:
                        return account_str
                    
                    level = get_account_level(account_str, lookup)
                    indent = "   " * level
                    name = lookup[account_str]['alias']
                    
                    if account_str.isdigit() and not name.startswith(account_str):
                        formatted_name = f"{account_str} {name}"
                    else:
                        formatted_name = name
                    
                    return f"{indent}{formatted_name}"
                
                # Create workbook
                wb = openpyxl.Workbook()
                wb.remove(wb.active)
                
                # Guidelines sheet
                guidelines = wb.create_sheet("Guidelines", 0)
                guidelines['A1'] = 'Company name'
                guidelines['B1'] = company_name
                guidelines['A2'] = 'Budget name'
                guidelines['B2'] = budget_name
                guidelines['A3'] = 'Budget type'
                guidelines['B3'] = 'Profit and loss'
                guidelines['A4'] = 'Vena Scenario'
                guidelines['B4'] = selected_scenario
                guidelines['A5'] = 'Year'
                guidelines['B5'] = int(selected_year)
                guidelines['A6'] = 'Period'
                guidelines['B6'] = f'1 - 12 (Jan {int(selected_year)} - Dec {int(selected_year)})'
                guidelines['A7'] = 'Subdivided by'
                guidelines['B7'] = 'Sub-Departments'
                
                # Get departments
                departments = sorted([str(d) for d in df['_Department'].unique()])
                
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                month_order = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                               'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
                
                for idx, dept in enumerate(departments):
                    status_text.text(f"Creating sheet for {dept}...")
                    
                    dept_data = df[df['_Department'] == dept].copy()
                    
                    pivot = dept_data.pivot_table(
                        index='_Account',
                        columns='Period_Name',
                        values='_value',
                        aggfunc='sum',
                        fill_value=0
                    )
                    
                    for month in month_order:
                        if month not in pivot.columns:
                            pivot[month] = 0
                    
                    pivot = pivot[month_order]
                    pivot['Budget totals'] = pivot.sum(axis=1)
                    pivot = pivot[['Budget totals'] + month_order]
                    
                    ws = wb.create_sheet(dept)
                    ws['A1'] = dept
                    ws['A1'].font = Font(bold=True)
                    
                    # Set column widths
                    ws.column_dimensions['A'].width = 52.8
                    ws.column_dimensions['B'].width = 15
                    for i in range(3, 15):
                        ws.column_dimensions[get_column_letter(i)].width = 13
                    
                    ws['A2'] = 'Accounts'
                    ws['B2'] = 'Budget totals'
                    for i, month in enumerate(month_order):
                        ws.cell(row=2, column=3+i, value=f'{month} {int(selected_year)}')
                    
                    current_row = 3
                    for account in pivot.index:
                        account_name = format_account_name(account, account_lookup)
                        ws.cell(row=current_row, column=1, value=account_name)
                        
                        for col_idx, col_name in enumerate(['Budget totals'] + month_order):
                            value = pivot.loc[account, col_name]
                            if value != 0:
                                ws.cell(row=current_row, column=2+col_idx, value=value)
                        
                        current_row += 1
                    
                    progress_bar.progress((idx + 1) / len(departments))
                
                status_text.text("✓ Complete!")
                
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                progress_bar.progress(1.0)
                
                st.success(f"▸ Generated QuickBooks budget file with {len(departments)} department tabs!")
                
                st.download_button(
                    label="↓ Download QuickBooks Budget File",
                    data=output.getvalue(),
                    file_name=f"QB_Upload_{selected_scenario}_{selected_year}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        except Exception as e:
            st.error(f"❌ **Error generating file:**\n\n```\n{str(e)}\n```")
            
else:
    st.info("↑ Please upload both CSV files to get started")

